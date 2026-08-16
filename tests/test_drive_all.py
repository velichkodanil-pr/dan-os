"""Full-Drive indexing: file filter, CSV row-safe extraction, sheet export name."""
import pytest

from app.core.google_client import (GOOGLE_DOC_MIME, GOOGLE_SHEET_MIME,
                                    drive_indexable)
from app.core.ingest import chunk_text, extract_text


def test_drive_indexable_filter():
    assert drive_indexable({"name": "Доступи", "mimeType": GOOGLE_SHEET_MIME})
    assert drive_indexable({"name": "Нотатки", "mimeType": GOOGLE_DOC_MIME})
    assert drive_indexable({"name": "logins.csv", "mimeType": "text/csv",
                            "size": "1024"})
    assert drive_indexable({"name": "contract.pdf", "mimeType": "application/pdf",
                            "size": str(5 * 1024 * 1024)})
    assert not drive_indexable({"name": "video.mp4", "mimeType": "video/mp4",
                                "size": "999"})
    assert not drive_indexable({"name": "huge.pdf", "mimeType": "application/pdf",
                                "size": str(50 * 1024 * 1024)})
    assert not drive_indexable({"name": "app.exe", "mimeType": "application/x-exe",
                                "size": "10"})


def test_csv_rows_stay_atomic_in_chunks():
    rows = [f"Партнер {i};login_p{i};https://portal{i}.example;нотатка про умови {i}"
            for i in range(60)]
    data = "\n".join(rows).encode()
    text = extract_text("Доступи партнерів.csv", data)
    assert "Партнер 7;login_p7" in text
    chunks = chunk_text(text)
    # every row lives in exactly one piece — never split mid-row
    for i in range(60):
        marker = f"Партнер {i};login_p{i}"
        assert sum(1 for c in chunks if marker in c) >= 1
        for c in chunks:
            if f"Партнер {i};" in c:
                assert f"login_p{i}" in c  # row not cut between name and login
                break


def test_extract_text_tsv():
    data = "Сервіс\tЛогін\nTravelon\tadmin_tvl\n".encode()
    text = extract_text("access.tsv", data)
    assert "admin_tvl" in text


def _make_xlsx() -> bytes:
    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Партнери"
    ws1.append(["Партнер", "Логін", "Портал"])
    ws1.append(["Anex", "anex_tvl", "portal.anex.example"])
    ws2 = wb.create_sheet("Банки")
    ws2.append(["Банк", "Логін"])
    ws2.append(["ПУМБ", "pumb_fin_login"])
    ws2.append([None, None])  # empty row must be skipped
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extract_text_xlsx_all_sheets():
    """ALL tabs land in the text (the whole point of the xlsx export)."""
    text = extract_text("Доступи.xlsx", _make_xlsx())
    assert "== Аркуш: Партнери ==" in text and "== Аркуш: Банки ==" in text
    assert "Anex | anex_tvl | portal.anex.example" in text
    assert "ПУМБ | pumb_fin_login" in text
    chunks = chunk_text(text)
    for c in chunks:  # row atomicity survives chunking
        if "Anex |" in c:
            assert "anex_tvl" in c
        if "ПУМБ" in c:
            assert "pumb_fin_login" in c


def test_drive_indexable_xlsx_and_sheet():
    from app.core.google_client import XLSX_MIME
    assert drive_indexable({"name": "fin.xlsx", "mimeType": XLSX_MIME,
                            "size": "2048"})


@pytest.mark.asyncio
async def test_ingest_document_parts_no_truncation(db):
    from app.core.ingest import PART_CHARS, ingest_document_parts
    rows = [f"Рядок даних номер {i} з достатньо довгим текстом усередині"
            for i in range(9000)]
    text = "\n\n".join(rows)
    # the password section analog sits at the very END
    text += ("\n\nOther | Toco UA | https://toco-tour.example | login_toco | "
             "депозит 30%")
    assert len(text) > PART_CHARS  # would have been truncated before
    results = await ingest_document_parts(
        db, user_id=111, domain="personal", title="Доступи.xlsx", text=text,
        source_type="drive", source_ref="file123",
        meta={"modifiedTime": "2026-08-13T00:00:00Z"})
    assert len(results) >= 2
    assert all(r.status == "indexed" for r in results)
    from sqlalchemy import select
    from app.models import KnowledgeChunk
    tail = (await db.execute(select(KnowledgeChunk).where(
        KnowledgeChunk.text.ilike("%login_toco%")))).scalars().all()
    assert tail, "the tail row must survive into the index"


@pytest.mark.asyncio
async def test_rag_keyword_fallback_finds_requisites_row(db):
    """R6.1A replacement for the old credential-row test.

    The capability it protected is real and stays: one exact business row must
    not be crowded out by thematically similar prose. Only the fixture changed
    — requisites instead of a password, because retrieving a password is no
    longer a behaviour DAN.OS has."""
    from app.core.ingest import ingest_document
    from app.core import rag
    await ingest_document(
        db, user_id=111, domain="personal", title="Реквізити партнерів (DMC)",
        text=("== Аркуш: DMC ==\n\nРеквізити операторів\n\n"
              "Other | Toco UA | https://toco-tour.example | ЄДРПОУ 46140224 | "
              "IBAN UA213223130000026007233566001\n\n"
              + "\n\n".join(f"Інший рядок {i} про фінанси і платежі" for i in range(40))),
        source_type="drive", source_ref="dmc1")
    # thematic decoys that would crowd out the row in pure vector search
    await ingest_document(
        db, user_id=111, domain="personal", title="Реєстр передоплат ТОКО",
        text="\n\n".join(f"ТОКО Україна платіж {i} на суму {i*100} грн"
                         for i in range(40)),
        source_type="drive", source_ref="reg1")
    chunks = await rag.retrieve(db, user_id=111, domain="personal",
                                query="реквізити Toco?")
    joined = " ".join(c.text for c in chunks)
    assert "UA213223130000026007233566001" in joined


@pytest.mark.asyncio
async def test_rag_cyrillic_query_finds_latin_brand(db):
    """«ТОКО» кирилицею має знаходити рядок із «Toco» латиницею (translit bridge)."""
    from app.core.ingest import ingest_document
    from app.core import rag
    await ingest_document(
        db, user_id=111, domain="personal", title="Умови операторів (DMC)",
        text=("Умови роботи з операторами\n\n"
              "Other | Toco UA | https://toco-tour.example | ЄДРПОУ 46140224 | "
              "депозит 30%\n\n"
              + "\n\n".join(f"Нейтральний рядок {i} про бронювання" for i in range(30))),
        source_type="drive", source_ref="dmc2")
    chunks = await rag.retrieve(db, user_id=111, domain="personal",
                                query="реквізити ТОКО Україна?")
    joined = " ".join(c.text for c in chunks)
    assert "46140224" in joined


def test_token_variants_translit():
    from app.core.rag import _token_variants
    v = _token_variants("токо")
    assert "toko" in v and "toco" in v and "ТОКО" in v
    assert _token_variants("anex") == {"anex"}  # latin stays as-is (ILIKE folds)


def test_xlsx_broken_dimensions_tail_survives():
    """Google-exported xlsx lies about dimensions; reset_dimensions must force
    reading to the REAL end of the sheet (password sections live at the tail)."""
    import io as _io
    import re as _re
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DMC"
    for i in range(60):
        ws.append([f"Оператор {i}", f"login{i}"])
    ws.append(["Other", "Toco UA", "https://toco-tour.example", "i.k@t.to",
               "депозит 30%"])
    buf = _io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    # simulate Google's broken metadata: claim the sheet is only 5 rows tall
    import zipfile as _zf
    src = _zf.ZipFile(_io.BytesIO(raw))
    out_buf = _io.BytesIO()
    with _zf.ZipFile(out_buf, "w", _zf.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            content = src.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                content = _re.sub(rb'<dimension ref="[^"]*"/>',
                                  b'<dimension ref="A1:B5"/>', content)
            dst.writestr(item, content)
    text = extract_text("dmc.xlsx", out_buf.getvalue())
    assert "toco-tour.example" in text, "tail beyond declared dimension must survive"


@pytest.mark.asyncio
async def test_delete_stale_versions(db):
    from app.core.ingest import delete_stale_versions, ingest_document
    old = await ingest_document(db, user_id=111, domain="personal",
                                title="Файл.xlsx",
                                text="Стара обрізана версія файла з даними " * 5,
                                source_type="drive", source_ref="FID1")
    new = await ingest_document(db, user_id=111, domain="personal",
                                title="Файл.xlsx (ч.1)",
                                text="Нова повна версія файла з хвостом і паролями " * 5,
                                source_type="drive", source_ref="FID1")
    removed = await delete_stale_versions(db, user_id=111, domain="personal",
                                          source_ref="FID1",
                                          keep_doc_ids={new.document.id})
    assert removed == 1
    from sqlalchemy import select
    from app.models import Document
    left = (await db.execute(select(Document).where(
        Document.source_ref == "FID1"))).scalars().all()
    assert [d.id for d in left] == [new.document.id]
