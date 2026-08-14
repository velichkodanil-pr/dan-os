"""Knowledge ingestion: scan -> extract text -> chunk -> embed -> store.

Content is DATA: nothing in an ingested file can trigger actions (policy stays
in code). Dedupe by content hash — re-ingesting the same content is a no-op.

R6.1A: the secret scan runs BEFORE chunking and BEFORE the embedder, so a
credential-bearing file costs zero provider calls and leaves zero source text
behind. What survives is a metadata-only Document row in `quarantined` state:
Danylo can see that a file was rejected and why (categories, counts) without
the content being stored, indexed or retrievable.
"""
import hashlib
import io
import logging
import re
import zipfile
from dataclasses import dataclass

from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core import security
from app.core.audit import audit
from app.core.embeddings import get_embedder
from app.models import Document, KnowledgeChunk

logger = logging.getLogger(__name__)

INGEST_VERSION = 2  # bump when extraction/splitting logic changes:
# skipped-by-modifiedTime files are re-processed under the new pipeline
MAX_DOC_BYTES = 15 * 1024 * 1024
MAX_CHARS = 400_000
CHUNK_SIZE = 800
CHUNK_OVERLAP = 120

ALLOWED_EXT = {".pdf", ".docx", ".txt", ".md", ".vtt", ".srt", ".csv", ".tsv", ".xlsx"}


def xlsx_to_sheets(data: bytes) -> list[tuple[str, str]]:
    """Workbook -> [(sheet_title, text)]. Each SHEET becomes its own document
    downstream, so a mega-workbook can never crowd its later tabs out of the
    index. Rows stay atomic paragraphs, because in a business table the ROW is
    the fact (partner, site, terms, contact) and splitting it destroys meaning.

    Per-sheet ingest also gives the security gate the right granularity: one
    tab with credentials is quarantined on its own, the rest of the workbook
    still gets indexed."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets: list[tuple[str, str]] = []
    for ws in wb.worksheets:
        try:
            ws.reset_dimensions()  # Google exports lie about dimensions
        except Exception:
            pass
        rows: list[str] = []
        total = 0
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                line = " | ".join(cells)
                rows.append(line)
                total += len(line)
            if len(rows) >= 20000 or total > 1_500_000:
                break
        if rows:
            sheets.append((ws.title, "\n\n".join(rows)))
    wb.close()
    return sheets


def _xlsx_to_text(data: bytes) -> str:
    """Workbook -> text: every sheet titled, every ROW an atomic paragraph
    (contact/terms tables must never split between the name and its values).

    Google-exported xlsx often carries WRONG dimension metadata; read_only
    iter_rows silently stops at the declared bound and drops the sheet tail.
    reset_dimensions() forces a full scan to the real end of each sheet — the
    security scan needs the whole sheet as much as retrieval does."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    total = 0
    for ws in wb.worksheets:
        try:
            ws.reset_dimensions()  # trust actual cells, not declared bounds
        except Exception:
            pass
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                rows.append(" | ".join(cells))
                total += len(rows[-1])
            if len(rows) >= 20000 or total > 2_400_000:  # sanity caps
                break
        if rows:
            parts.append(f"== Аркуш: {ws.title} ==\n\n" + "\n\n".join(rows))
        if total > 2_400_000:
            break
    wb.close()
    return "\n\n".join(parts)


class IngestError(Exception):
    pass


_TS_RE = re.compile(r"-->")
_CUE_NUM_RE = re.compile(r"^\d+$")
_SPEAKER_RE = re.compile(r"^([^:\n]{1,60}):\s*(.+)$", re.DOTALL)


def parse_subtitles(data: bytes) -> str:
    """WEBVTT/SRT (Zoom meeting transcripts) -> clean 'Speaker: text' lines.

    Drops headers, cue numbers and timestamps; merges consecutive cues of the
    same speaker so the transcript reads as a dialogue, not subtitle spam."""
    raw = data.decode("utf-8", "ignore").replace("\r\n", "\n")
    lines: list[tuple[str, str]] = []  # (speaker, text)
    for line in raw.split("\n"):
        line = line.strip().lstrip("﻿")
        if (not line or line.upper().startswith(("WEBVTT", "NOTE", "STYLE"))
                or _TS_RE.search(line) or _CUE_NUM_RE.match(line)):
            continue
        m = _SPEAKER_RE.match(line)
        speaker, text = (m.group(1).strip(), m.group(2).strip()) if m else ("", line)
        if lines and lines[-1][0] == speaker:
            prev_speaker, prev_text = lines[-1]
            if text not in prev_text[-len(text) * 2:]:  # skip verbatim repeats
                lines[-1] = (prev_speaker, f"{prev_text} {text}")
        else:
            lines.append((speaker, text))
    out = [f"{s}: {t}" if s else t for s, t in lines]
    return "\n".join(out)


def extract_text(filename: str, data: bytes) -> str:
    name = filename.lower()
    if len(data) > MAX_DOC_BYTES:
        raise IngestError("Файл завеликий (ліміт 15 МБ)")
    if name.endswith((".vtt", ".srt")):
        text = parse_subtitles(data)
        text = re.sub(r"[ \t]+", " ", text).strip()
        if len(text) < 20:
            raise IngestError("У транскрипті не знайшлося тексту")
        return text[:MAX_CHARS]
    if name.endswith(".pdf"):
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            text = "\n".join((page.extract_text() or "") for page in reader.pages[:200])
        except Exception as e:
            raise IngestError("Не зміг прочитати PDF") from e
    elif name.endswith(".docx"):
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                xml = z.read("word/document.xml").decode("utf-8", "ignore")
            xml = re.sub(r"</w:p>", "\n", xml)
            text = re.sub(r"<[^>]+>", "", xml)
        except Exception as e:
            raise IngestError("Не зміг прочитати DOCX") from e
    elif name.endswith(".xlsx"):
        try:
            text = _xlsx_to_text(data)
        except Exception as e:
            raise IngestError("Не зміг прочитати XLSX") from e
        text = re.sub(r"[ \t]+", " ", text).strip()
        if len(text) < 20:
            raise IngestError("У таблиці не знайшлося тексту")
        return text[:2_000_000]  # big workbooks split into parts at ingest
    elif name.endswith((".csv", ".tsv")):
        raw = data.decode("utf-8", "ignore")
        # each row becomes a paragraph -> chunker never cuts a row in half
        # (critical for contact/terms tables: the row IS the fact)
        text = "\n\n".join(line for line in raw.splitlines() if line.strip())
    elif name.endswith((".txt", ".md")):
        text = data.decode("utf-8", "ignore")
    else:
        raise IngestError("Підтримую pdf, docx, txt, md, csv")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    if len(text) < 20:
        raise IngestError("У файлі не знайшлося тексту")
    return text[:MAX_CHARS]


def chunk_text(text: str, size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    chunks: list[str] = []
    current = ""
    for p in paragraphs:
        while len(p) > size:  # oversized paragraph — hard split
            if current:
                chunks.append(current)
                current = ""
            chunks.append(p[:size])
            p = p[size - overlap:]
        if len(current) + len(p) + 1 <= size:
            current = f"{current}\n{p}".strip()
        else:
            chunks.append(current)
            current = (current[-overlap:] + "\n" + p).strip() if overlap else p
    if current:
        chunks.append(current)
    return [c for c in chunks if len(c) > 30]


@dataclass
class IngestResult:
    status: str  # indexed | duplicate | quarantined | error
    document: Document | None = None
    chunks: int = 0
    error: str | None = None
    categories: tuple = ()  # secret categories when quarantined (never values)


async def ingest_document(
    db: AsyncSession, *, user_id: int, title: str, text: str,
    source_type: str, source_ref: str = "", domain: str = "personal",
    meta: dict | None = None,
) -> IngestResult:
    content_hash = hashlib.sha256(text.strip().lower().encode()).hexdigest()
    existing = (await db.execute(
        select(Document).where(Document.content_hash == content_hash))).scalar_one_or_none()
    if existing:
        await audit(db, actor=f"user:{user_id}", action="ingest", resource_type="document",
                    resource_id=existing.id, outcome="dedupe", title=title)
        await db.commit()
        # re-uploading a quarantined file stays quarantined and stays silent
        # about its content — the dedupe path must not become a read channel
        status = "quarantined" if existing.status == "quarantined" else "duplicate"
        return IngestResult(status=status, document=existing,
                            chunks=existing.chunk_count)

    # ---- the gate: nothing below this line may see a credential ----
    scan = security.scan(text)
    if scan.blocked:
        doc = Document(user_id=user_id, domain=domain, title=title[:200],
                       source_type=source_type, source_ref=source_ref[:500],
                       content_hash=content_hash, status="quarantined",
                       chunk_count=0,
                       meta={**(meta or {}), "security": scan.as_meta()})
        db.add(doc)
        try:
            await db.flush()
        except IntegrityError:  # race on hash
            await db.rollback()
            return IngestResult(status="duplicate")
        await security.record_finding(db, user_id=user_id, domain=domain,
                                      resource_type="document",
                                      resource_id=doc.id, result=scan)
        await security.audit_blocked(db, user_id=user_id, action="ingest.quarantined",
                                     resource_type="document", resource_id=doc.id,
                                     result=scan)
        await db.commit()
        logger.info("ingest quarantined: source=%s categories=%s findings=%d",
                    source_type, ",".join(str(c) for c in scan.categories),
                    scan.finding_count)
        return IngestResult(status="quarantined", document=doc, chunks=0,
                            categories=scan.categories)

    chunks = chunk_text(text)
    if not chunks:
        return IngestResult(status="error", error="Не вийшло розбити текст")
    embeddings = await get_embedder().embed(chunks)

    doc = Document(user_id=user_id, domain=domain, title=title[:200],
                   source_type=source_type, source_ref=source_ref[:500],
                   content_hash=content_hash, chunk_count=len(chunks),
                   meta=meta or {})
    db.add(doc)
    try:
        await db.flush()
    except IntegrityError:  # race on hash
        await db.rollback()
        return IngestResult(status="duplicate")
    for i, (chunk, emb) in enumerate(zip(chunks, embeddings)):
        db.add(KnowledgeChunk(document_id=doc.id, user_id=user_id, seq=i,
                              text=chunk, embedding=emb))
    await audit(db, actor=f"user:{user_id}", action="ingest", resource_type="document",
                resource_id=doc.id, policy_level="L1", title=title, chunks=len(chunks))
    await db.commit()
    return IngestResult(status="indexed", document=doc, chunks=len(chunks))


async def delete_stale_versions(db: AsyncSession, *, user_id: int,
                                source_ref: str, keep_doc_ids: set) -> int:
    """Drop older ingested versions of the same Drive file (truncated extracts,
    first-tab csv exports) so retrieval sees ONE current version. Chunks go via
    FK cascade. Returns how many stale documents were removed."""
    if not source_ref:
        return 0
    stale = (await db.execute(
        select(Document).where(Document.user_id == user_id,
                               Document.source_type == "drive",
                               Document.source_ref == source_ref))).scalars().all()
    removed = 0
    for doc in stale:
        if doc.id not in keep_doc_ids:
            await db.delete(doc)
            removed += 1
    if removed:
        await audit(db, actor=f"user:{user_id}", action="ingest.stale_removed",
                    resource_type="document", resource_id=source_ref[:60],
                    removed=removed)
    await db.commit()
    return removed


async def ingest_xlsx_by_sheets(
    db: AsyncSession, *, user_id: int, filename: str, data: bytes,
    source_type: str, source_ref: str = "", domain: str = "personal",
    meta: dict | None = None,
) -> list[IngestResult]:
    """Every sheet -> its own document «файл · аркуш "Назва"» (parts if huge)."""
    try:
        sheets = xlsx_to_sheets(data)
    except Exception as e:
        raise IngestError("Не зміг прочитати XLSX") from e
    if not sheets:
        raise IngestError("У таблиці не знайшлося тексту")
    base = filename.rsplit(".", 1)[0][:120]
    results: list[IngestResult] = []
    for sheet_title, text in sheets:
        text = re.sub(r"[ \t]+", " ", text).strip()
        if len(text) < 20:
            continue
        title = f"{base} · аркуш «{sheet_title[:40]}»"
        results.extend(await ingest_document_parts(
            db, user_id=user_id, title=title, text=text,
            source_type=source_type, source_ref=source_ref,
            domain=domain, meta=meta))
    return results


PART_CHARS = 350_000


async def ingest_document_parts(
    db: AsyncSession, *, user_id: int, title: str, text: str,
    source_type: str, source_ref: str = "", domain: str = "personal",
    meta: dict | None = None,
) -> list[IngestResult]:
    """Oversized texts (huge workbooks/exports) become «title (ч.N)» parts.

    The scan runs on the WHOLE text before splitting, so a credential block
    cannot be smuggled in by landing on a part boundary; the whole source is
    then contained as one metadata-only row rather than partly indexed.
    """
    scan = security.scan(text)
    if scan.blocked:
        return [await ingest_document(
            db, user_id=user_id, title=title, text=text, source_type=source_type,
            source_ref=source_ref, domain=domain, meta=meta)]
    if len(text) <= PART_CHARS:
        return [await ingest_document(db, user_id=user_id, title=title, text=text,
                                      source_type=source_type, source_ref=source_ref,
                                      domain=domain, meta=meta)]
    paragraphs = text.split("\n\n")
    parts: list[str] = []
    current: list[str] = []
    size = 0
    for para in paragraphs:
        if size + len(para) > PART_CHARS and current:
            parts.append("\n\n".join(current))
            current, size = [], 0
        current.append(para)
        size += len(para) + 2
    if current:
        parts.append("\n\n".join(current))
    results = []
    for i, part in enumerate(parts, 1):
        results.append(await ingest_document(
            db, user_id=user_id, title=f"{title} (ч.{i})" if len(parts) > 1 else title,
            text=part, source_type=source_type, source_ref=source_ref,
            domain=domain, meta=meta))
    return results
